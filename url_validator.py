import asyncio
import aiohttp
import pandas as pd
import argparse
import time
import re
from tenacity import retry, stop_after_attempt, wait_exponential
from urllib.parse import urlparse
import openpyxl

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
async def validate_url(session, url, semaphore, rate_limit_delay=0.1):
    async with semaphore:
        await asyncio.sleep(rate_limit_delay)  # Rate limiting
        start_time = time.time()
        try:
            # Try HEAD request first
            async with session.head(url, timeout=aiohttp.ClientTimeout(total=10)) as response:
                status = response.status
                redirect = str(response.url) if str(response.url) != url else None
                response_time = time.time() - start_time
                return {
                    'url': url,
                    'status': 'valid' if status < 400 else 'invalid',
                    'http_code': status,
                    'response_time': round(response_time, 2),
                    'redirect': redirect,
                    'error': None
                }
        except Exception:
            try:
                # Fallback to GET
                async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as response:
                    status = response.status
                    redirect = str(response.url) if str(response.url) != url else None
                    response_time = time.time() - start_time
                    return {
                        'url': url,
                        'status': 'valid' if status < 400 else 'invalid',
                        'http_code': status,
                        'response_time': round(response_time, 2),
                        'redirect': redirect,
                        'error': None
                    }
            except Exception as e:
                response_time = time.time() - start_time
                return {
                    'url': url,
                    'status': 'invalid',
                    'http_code': None,
                    'response_time': round(response_time, 2),
                    'redirect': None,
                    'error': str(e)
                }

def get_verbose_error_explanation(status, http_code, error):
    """Generate a verbose explanation for URL validation failures."""
    if status == 'error':
        if 'timeout' in error.lower():
            return f"Connection timeout: The server did not respond within the allowed time (10 seconds). This could indicate server overload, network issues, or the server is down."
        elif 'connection' in error.lower():
            return f"Connection failed: Unable to establish a connection to the server. Possible causes include DNS resolution failure, firewall blocking, or the server being unreachable."
        elif 'ssl' in error.lower():
            return f"SSL/TLS error: Secure connection could not be established. This might be due to an invalid or expired SSL certificate, or SSL configuration issues."
        else:
            return f"Request error: {error}. This indicates a problem with the HTTP request that prevented validation."
    
    elif status == 'invalid' and http_code:
        if http_code == 404:
            return f"Page not found (404): The requested URL does not exist on the server. The page may have been moved, deleted, or the URL is incorrect."
        elif http_code == 403:
            return f"Forbidden (403): Access to the URL is denied. This could be due to authentication requirements, IP blocking, or server configuration."
        elif http_code == 500:
            return f"Internal server error (500): The server encountered an unexpected condition that prevented it from fulfilling the request. This is a server-side issue."
        elif http_code == 502:
            return f"Bad gateway (502): The server received an invalid response from an upstream server. This indicates a problem with the server's proxy or backend services."
        elif http_code == 503:
            return f"Service unavailable (503): The server is temporarily unable to handle the request, possibly due to maintenance or overload."
        elif http_code >= 400 and http_code < 500:
            return f"Client error ({http_code}): The request contains bad syntax or cannot be fulfilled. This suggests an issue with the URL format or request parameters."
        elif http_code >= 500:
            return f"Server error ({http_code}): The server failed to fulfill a valid request. This indicates a problem on the server's side."
        else:
            return f"Invalid response ({http_code}): The server returned an unexpected status code that indicates the URL is not accessible."
    
    return error or "Unknown validation failure"

async def main():
    parser = argparse.ArgumentParser(description='Validate URLs from an Excel file by scanning all cells.')
    parser.add_argument('input_file', help='Path to the input Excel file (.xlsx)')
    parser.add_argument('--output', default='validated_urls.xlsx', help='Path to the output Excel file (default: validated_urls.xlsx)')
    parser.add_argument('--concurrency', type=int, default=10, help='Maximum number of concurrent requests (default: 10)')
    parser.add_argument('--rate_limit', type=float, default=0.1, help='Delay between requests in seconds (default: 0.1)')

    args = parser.parse_args()

    # URL regex pattern
    url_pattern = re.compile(r'https?://[^\s]+')

    # Read input file and extract URLs from all cells
    urls_with_locations = []
    try:
        wb = openpyxl.load_workbook(args.input_file, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        # Find all URLs in the cell value
                        matches = url_pattern.findall(cell.value)
                        for url in matches:
                            urls_with_locations.append({
                                'url': url,
                                'sheet': sheet_name,
                                'row': row,
                                'column': col,
                                'cell_value': cell.value
                            })
    except Exception as e:
        print(f"Error reading input file: {e}")
        return

    if not urls_with_locations:
        print("No URLs found in the input file.")
        return

    urls = [item['url'] for item in urls_with_locations]

    print(f"Found {len(urls)} URLs to validate.")

    semaphore = asyncio.Semaphore(args.concurrency)

    async with aiohttp.ClientSession() as session:
        tasks = [validate_url(session, url, semaphore, args.rate_limit) for url in urls]
        results = await asyncio.gather(*tasks, return_exceptions=True)

    # Handle exceptions in results
    processed_results = []
    for i, result in enumerate(results):
        location = urls_with_locations[i]
        if isinstance(result, Exception):
            processed_results.append({
                'url': location['url'],
                'sheet': location['sheet'],
                'row': location['row'],
                'column': location['column'],
                'cell_value': location['cell_value'],
                'status': 'error',
                'http_code': None,
                'response_time': None,
                'redirect': None,
                'error': str(result)
            })
        else:
            processed_results.append({
                'url': location['url'],
                'sheet': location['sheet'],
                'row': location['row'],
                'column': location['column'],
                'cell_value': location['cell_value'],
                **result
            })

    # Enhance error explanations and sort results
    for item in processed_results:
        if item['status'] != 'valid':
            item['error'] = get_verbose_error_explanation(item['status'], item['http_code'], item['error'])

    # Sort: failed URLs first, then valid ones
    def sort_key(item):
        status_order = {'error': 0, 'invalid': 1, 'valid': 2}
        return status_order.get(item['status'], 3)
    
    processed_results.sort(key=sort_key)

    # Create output DataFrame
    output_df = pd.DataFrame(processed_results)

    # Save to output file
    try:
        output_df.to_excel(args.output, index=False)
        print(f"Validation complete. Results saved to {args.output}")
    except Exception as e:
        print(f"Error saving output file: {e}")

if __name__ == '__main__':
    asyncio.run(main())